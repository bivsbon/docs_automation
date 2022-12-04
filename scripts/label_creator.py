import openpyxl as xl


def create_label(path, start_index=1):
    wb_obj = xl.load_workbook(path)
    sheet_obj = wb_obj.active

    with open('label2.txt', 'w', encoding='utf-8') as f:
        for i in range(2, sheet_obj.max_row + 1):
            cell_obj = sheet_obj.cell(row=i, column=27)
            f.write('dat/%d.jpg\t%s\n' % (i-2+start_index, str(cell_obj.value)))


create_label("phuonghaichau2_qhc_tpdn.KH.2007.01.xlsx", start_index=898)
