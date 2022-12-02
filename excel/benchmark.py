import openpyxl as xl

# workbook = xlrd.open_workbook(r"excel\phuonghaichau2_qhc_tpdn.KH.2008.02.xls")
# sheet = workbook.sheet_by_name("Sheet")
# #getting the first sheet
# sheet_1 = workbook.sheet_by_index(0)
#
# for sh in workbook.sheets():
#     print(sh.name)
#
# row_count = sheet.nrows
# col_count = sheet.ncols
#
# for cur_row in range(0, row_count):
#     for cur_col in range(0, col_count):
#         cell = sheet.cell(cur_row, cur_col)
#         print(cell.value, cell.ctype, end='')
#     print()
path = "excel\\phuonghaichau2_qhc_tpdn.KH.2008.02.xlsx"

# To open the workbook
# workbook object is created
wb_obj = xl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# Cell objects also have a row, column,
# and coordinate attributes that provide
# location information for the cell.

# Note: The first row or
# column integer is 1, not 0.

# Cell object is created by using
# sheet object's cell() method.

with open('label.txt', 'w', encoding='utf-8') as f:
    for i in range(2, sheet_obj.max_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=16)
        f.write('sample/%d.jpg\t%s\n' % (i-1, str(cell_obj.value)))
