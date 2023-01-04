from concurrent.futures import ThreadPoolExecutor

import openpyxl as xl
import os
from img_processing_module import ImgProcessingModule
from file_mapper import get_mappings
from model.model import TextRecognizer
from openpyxl.styles import PatternFill
from no_accent_vietnamese import convert

crit_cols_conf = {
    1: [[16, 27], [17, 28]]
}

field_col_maps = {
    1: {16: 0, 17: 1, 27: 2, 28: 3}
}


def different(s1, s2):
    if len(s1) != len(s2):
        return True
    else:
        count = 0
        for i in range(len(s1)):
            if s1[i] != s2[i]:
                count += 1
            if count == 2:
                return True
    return False


def fill_value_in(cell, value, max_row):
    print(f"{cell.row - 1}/{max_row - 1}")
    print(f"Changing {str(cell.value)} into {value} in coord {cell.coordinate}")
    cell.value = value
    cell.fill = PatternFill(start_color="FFE900", end_color="FFE900", fill_type="solid")


def process_img(img_dir, i, sheet, mode, img_module, recognizers):
    global crit_cols_conf
    global field_col_maps
    crit_cols = crit_cols_conf[mode]
    field_col_map = field_col_maps[mode]
    result = []

    img_path = f'{img_dir}\\{i - 1}.jpg'
    fields = img_module.get_fields(img_path)

    # Check if number of detected fields match number of columns
    if len(fields) == len(field_col_map):
        for gi in range(len(crit_cols)):
            # Each column group uses different model for text recognition
            col_group = crit_cols[gi]
            for col_index in col_group:
                field_index = field_col_map[col_index]
                prediction = recognizers[gi].predict(fields[field_index])
                cell = sheet.cell(row=i, column=col_index)

                # Check if prediction is the same as value in Excel
                # if different(str(cell.value), prediction):
                if convert(prediction) != convert(str(cell.value)):
                    result.append([cell, prediction])
                    # Fill the value in color that cell yellow
                    # fill_value_in(cell, prediction, sheet.max_row)

                # else:
                #     print(f'skip{cell.row-1}/{sheet.max_row-1}')
    else:
        # TODO: Field detection error
        print(f"Field detection error at index {i}")

    return result


def main():
    img_dir = 'img'
    excel_dir = 'excel'

    recognizers = [TextRecognizer('./weights/name.pth'), TextRecognizer('weights/date.pth')]

    # TODO: find some ways to input mode
    mode = 1
    img_module = ImgProcessingModule(mode)

    # Get mapping between image directory and Excel file
    mappings = get_mappings()
    # mapping[0] is img directory, mapping[1] is Excel file
    for mapping in mappings:
        # Load Excel sheet and create essential variables
        path = os.path.join(excel_dir, mapping[1])
        wb = xl.load_workbook(path)
        sheet = wb.active
        max_row = sheet.max_row
        img_dir2 = os.path.join(img_dir, mapping[0])
        executor = ThreadPoolExecutor(max_workers=5)
        futures = []

        # Loop through sheet rows and images
        # for i in range(2, 7):
        for i in range(2, sheet.max_row + 1):
            future = executor.submit(process_img, img_dir2, i, sheet, mode, img_module, recognizers)
            futures.append(future)

        for future in futures:
            result = future.result()
            for insert in result:
                fill_value_in(insert[0], insert[1], sheet.max_row)

        wb.save("modified.xlsx")


if __name__ == "__main__":
    main()
